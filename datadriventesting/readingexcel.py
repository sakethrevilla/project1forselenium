# import openpyxl
#
# # file/worrkbook/sheeets/rows/cells
#
# file="F:/selenium1.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook["Sheet2"]
#
# rows=sheet.max_row
# columnes=sheet.max_column
#
# # redaing the all the data from the sheets
#
# for r in range(1,(rows+1)):
#     for c in range(1,(columnes+1)):
#         print(sheet.cell(r,c).value,end="  ")
#     print()
# print(rows)


import openpyxl

file='F:/writing2.xlsx'
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row
columns=sheet.max_column

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value ,end='  ')
    print()
    print()