import openpyxl

file="F:\writing.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
# here we are writing the same values
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="saketh"
workbook.save(file)

# writing the different values


file="F:\writing.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook.active

sheet.cell(1,1).value="name"
sheet.cell(1,2).value="age"
sheet.cell(1,3).value="country"

sheet.cell(2,1).value="saketh"
sheet.cell(2,2).value="22"
sheet.cell(2,3).value="india"

sheet.cell(3,1).value="sam"
sheet.cell(3,2).value="25"
sheet.cell(3,3).value="us"

sheet.cell(4,1).value="ram"
sheet.cell(4,2).value="20"
sheet.cell(4,3).value="uk"
workbook.save(file)

