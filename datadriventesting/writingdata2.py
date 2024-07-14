import openpyxl

# file="F:/writing.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook.active
# # here we are printing the same data for 4 rows with six columns
# for r in range(1,4):
#     for c in range(1,6):
#         sheet.cell(r,c).value="india"
# workbook.save(file)

# method2
# we are writing the data seperately across the multiple tables

file="F:/writing2.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active

sheet.cell(1,1).value="Name"
sheet.cell(1,2).value="Age"
sheet.cell(1,3).value="salary"

sheet.cell(2,1).value="rama"
sheet.cell(2,2).value=10
sheet.cell(2,3).value="365821"

sheet.cell(3,1).value="gopi"
sheet.cell(3,2).value=18
sheet.cell(3,3).value="35821"

workbook.save(file)
